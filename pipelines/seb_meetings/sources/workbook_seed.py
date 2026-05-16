"""Workbook → warehouse seed loader.

Reads the v0 dashboard workbook (`fixtures/seb_meetings_v0.xlsx`) and emits
typed records for the three SEB tables: meetings, controversies, sources.

Why this exists
---------------
Before live RSS ingestion came online, all SEB meeting tracking happened
in a single spreadsheet. The pipeline cannot pretend that history did not
exist — it must absorb the workbook as its starting point. The workbook
is therefore treated as a first-class source, on equal footing with the
YouTube RSS feed, with one important difference: it loads exactly once,
not on every flow run.

Design notes
------------
- Source = "workbook_v0". Anything loaded here is forever attributable.
- Meeting IDs come from the workbook's `#` column. They are stable.
- Hyperlinks are extracted from openpyxl cells (`cell.hyperlink.target`),
  not from displayed text — the workbook uses display labels like
  "Watch" and "Source" instead of bare URLs.
- This file does NOT correct data quality issues found in the workbook.
  If a row's video URL points to the wrong meeting (real example: see
  LESSONS.md §3), that lives in the warehouse as-is. Corrections happen
  in a separate, auditable migration with provenance — never silently
  rewritten on ingest.
- The Pydantic `Meeting` model is the validator. If a row fails schema,
  the loader fails loudly with the row number and field — it never
  silently coerces.
"""

from __future__ import annotations

import datetime as dt
from collections.abc import Iterator
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from pydantic import ValidationError

from pipelines.seb_meetings.transforms.models import (
    ComplianceStatus,
    Meeting,
    MeetingType,
)

# ---------------------------------------------------------------------------
# Constants — anchored to the v0 workbook structure. If the workbook layout
# changes, change these here, not at call sites.
# ---------------------------------------------------------------------------

MEETINGS_SHEET = "Meetings"
CONTROVERSIES_SHEET = "Controversies"
SOURCES_SHEET = "Sources"

# Row 5 is the column header band in all three sheets. Data starts at row 6.
HEADER_ROW = 5
DATA_START_ROW = 6

# Compliance column displays "Compliant" / "Notice Concerns" / etc. but the
# v0 workbook also has blank cells where review never happened.
_COMPLIANCE_BLANK_FALLBACK = ComplianceStatus.UNREVIEWED


# Display labels the workbook uses for link cells. When a cell shows one
# of these but has no hyperlink target, the URL is genuinely missing —
# returning the literal label would corrupt the warehouse.
_LINK_DISPLAY_LABELS = {"Watch", "Source", "Link", "Notes", ""}


def _hyperlink_or_text(cell: Cell) -> str | None:
    """Return the cell's hyperlink target if present, else its text value.

    The workbook uses display labels like "Watch" / "Source" with hyperlink
    targets, so the cell's `.value` is useless on its own. If the cell has
    no hyperlink AND its text matches a known link-label, treat the URL as
    missing — these cells exist in the workbook even where no URL was
    available, which is itself a real data-quality signal.
    """
    if cell.hyperlink is not None and cell.hyperlink.target:
        return cell.hyperlink.target
    val = cell.value
    if val is None:
        return None
    s = str(val).strip()
    if not s or s in _LINK_DISPLAY_LABELS:
        return None
    # Only return raw text if it actually looks like a URL.
    if s.startswith(("http://", "https://")):
        return s
    return None


def _day_abbrev(value: object) -> str:
    """Normalize 'Thu', 'Thursday', 'thu ' → 'Thu' style 3-char abbrev."""
    if value is None:
        return ""
    raw = str(value).strip()
    # Workbook already uses 3-char abbreviations — preserve them.
    return raw[:3]


def _parse_date(value: object, row: int) -> dt.date:
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    raise ValueError(f"Meetings row {row}: expected a date in column 'Date', got {value!r}")


def _parse_compliance(value: object) -> ComplianceStatus:
    if value is None or (isinstance(value, str) and not value.strip()):
        return _COMPLIANCE_BLANK_FALLBACK
    raw = str(value).strip()
    try:
        return ComplianceStatus(raw)
    except ValueError as err:
        raise ValueError(
            f"Unknown compliance status {raw!r}. Allowed: {[c.value for c in ComplianceStatus]}"
        ) from err


def _parse_meeting_type(value: object, row: int) -> MeetingType:
    if value is None:
        raise ValueError(f"Meetings row {row}: meeting_type is required")
    raw = str(value).strip()
    try:
        return MeetingType(raw)
    except ValueError as err:
        raise ValueError(
            f"Meetings row {row}: unknown meeting_type {raw!r}. "
            f"Allowed: {[t.value for t in MeetingType]}"
        ) from err


def _parse_hours(value: object, row: int) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError) as err:
        raise ValueError(
            f"Meetings row {row}: hours_logged must be numeric, got {value!r}"
        ) from err


def iter_meetings(workbook_path: Path) -> Iterator[Meeting]:
    """Yield validated `Meeting` instances from the workbook's Meetings sheet.

    The workbook is opened with `data_only=False` so cell hyperlinks are
    preserved. Closed at the end of iteration.
    """
    wb = load_workbook(workbook_path, data_only=False)
    try:
        ws = wb[MEETINGS_SHEET]
        for row in range(DATA_START_ROW, ws.max_row + 1):
            meeting_id = ws.cell(row, 2).value
            if meeting_id is None:
                continue  # blank row guard
            try:
                meeting = Meeting(
                    meeting_id=int(meeting_id),
                    meeting_date=_parse_date(ws.cell(row, 3).value, row),
                    day_of_week=_day_abbrev(ws.cell(row, 4).value),
                    meeting_type=_parse_meeting_type(ws.cell(row, 5).value, row),
                    meeting_format=str(ws.cell(row, 6).value or "").strip(),
                    chair=str(ws.cell(row, 7).value or "").strip(),
                    members_present=str(ws.cell(row, 8).value or "").strip(),
                    quorum_met=str(ws.cell(row, 9).value or "").strip(),
                    agenda_summary=ws.cell(row, 10).value or None,
                    key_decisions=ws.cell(row, 11).value or None,
                    video_url=_hyperlink_or_text(ws.cell(row, 12)),
                    source_url=_hyperlink_or_text(ws.cell(row, 13)),
                    compliance_status=_parse_compliance(ws.cell(row, 14).value),
                    compliance_notes=ws.cell(row, 15).value or None,
                    controversies=ws.cell(row, 16).value or None,
                    hours_logged=_parse_hours(ws.cell(row, 17).value, row),
                )
            except ValidationError as err:
                # Surface the row number so course exercises can debug.
                raise ValueError(f"Meetings row {row} failed Pydantic validation: {err}") from err
            yield meeting
    finally:
        wb.close()


def iter_controversies(workbook_path: Path) -> Iterator[dict]:
    """Yield raw controversy dicts from the workbook.

    Controversies are still untyped at the warehouse layer in Phase 1 — see
    `warehouse/schema/seb.sql` and ADR-0001. When we promote them to a
    Pydantic model in Phase 2, this function gains a typed return.
    """
    wb = load_workbook(workbook_path, data_only=False)
    try:
        ws = wb[CONTROVERSIES_SHEET]
        for row in range(DATA_START_ROW, ws.max_row + 1):
            cid = ws.cell(row, 2).value
            if cid is None:
                continue
            first_seen_raw = ws.cell(row, 4).value
            yield {
                "controversy_id": int(cid),
                "title": str(ws.cell(row, 3).value).strip(),
                "first_seen_date": _coerce_loose_date(first_seen_raw),
                "status": str(ws.cell(row, 5).value or "Active").strip(),
                "description": ws.cell(row, 6).value,
                "latest_action": ws.cell(row, 7).value,
                "primary_source": _hyperlink_or_text(ws.cell(row, 8)),
            }
    finally:
        wb.close()


def iter_sources(workbook_path: Path) -> Iterator[dict]:
    """Yield raw source-citation dicts from the workbook."""
    wb = load_workbook(workbook_path, data_only=False)
    try:
        ws = wb[SOURCES_SHEET]
        for row in range(DATA_START_ROW, ws.max_row + 1):
            sid = ws.cell(row, 2).value
            if sid is None:
                continue
            url_cell = ws.cell(row, 5)
            yield {
                "source_id": int(sid),
                "name": str(ws.cell(row, 3).value).strip(),
                "source_type": str(ws.cell(row, 4).value or "Other").strip(),
                "url": _hyperlink_or_text(url_cell) or "",
                "notes": ws.cell(row, 6).value,
            }
    finally:
        wb.close()


def _coerce_loose_date(value: object) -> dt.date | None:
    """The Controversies sheet uses 'YYYY-MM' strings, not real dates.

    We accept either a real date or a 'YYYY-MM' string and coerce to a
    first-of-month date. Anything else returns None and lets the warehouse
    NULL the column.
    """
    if value is None or value == "":
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    s = str(value).strip()
    # 'YYYY-MM'
    if len(s) == 7 and s[4] == "-":
        try:
            return dt.date(int(s[:4]), int(s[5:7]), 1)
        except ValueError:
            return None
    # 'YYYY-MM-DD'
    try:
        return dt.date.fromisoformat(s)
    except ValueError:
        return None


def load_all(workbook_path: Path) -> dict[str, list]:
    """Materialize all three tables from the workbook into Python lists.

    Returns a dict with keys 'meetings', 'controversies', 'sources'.
    Designed for use by both the Prefect flow and ad-hoc REPL exploration.
    """
    return {
        "meetings": list(iter_meetings(workbook_path)),
        "controversies": list(iter_controversies(workbook_path)),
        "sources": list(iter_sources(workbook_path)),
    }
